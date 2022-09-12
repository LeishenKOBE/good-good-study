import openpyxl

# wb = openpyxl.Workbook()

# sheet = wb.active

# sheet.title = 'new title'

# sheet['A1'] = '漫威宇宙'
# row = ['美国队长', '钢铁侠', '蜘蛛侠']

# sheet.append(row)

# wb.save('Marvel.xlsx')


wb = openpyxl.load_workbook('Marvel.xlsx')

sheet = wb['new title']
sheetname = wb.sheetnames

A1_cell = sheet['A1']
A1_value = A1_cell.value

print(A1_value)
